"""
Kernlogik des Excel Editors.
Lesen und Schreiben von Excel-Dateien mit Erhalt der Formatierung.

Verwendet openpyxl – kein laufendes Excel erforderlich.
OneDrive/SharePoint-AutoSync bleibt erhalten: der OneDrive-Client erkennt
die geänderte Datei und synchronisiert sie automatisch.
"""

from __future__ import annotations

import copy
import re
import shutil
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import PatternFill

from .models import CellInfo, ExcelReadConfig, RowData, SheetInfo


# ---------------------------------------------------------------------------
# Hilfsfunktionen (privat)
# ---------------------------------------------------------------------------

# Namespace-Deklarationen, die für Microsoft-365-Erweiterungen benötigt werden
_MS_EXT_NAMESPACES = [
    ("x15",  "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main"),
    ("x14",  "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"),
    ("xr",   "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"),
    ("xr2",  "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"),
    ("xr6",  "http://schemas.microsoft.com/office/spreadsheetml/2016/revision6"),
    ("xr10", "http://schemas.microsoft.com/office/spreadsheetml/2016/revision10"),
]


def _read_workbook_extlst(file_path: Path) -> Optional[str]:
    """
    Liest das <extLst>-Element aus xl/workbook.xml.
    Dieses Element enthält u.a. <x15:workbookPr autoSave="1"/>,
    das die AutoSave-Funktion in Excel steuert.
    """
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            if "xl/workbook.xml" not in zf.namelist():
                return None
            content = zf.read("xl/workbook.xml").decode("utf-8")
            match = re.search(r"<extLst\b[^>]*>.*?</extLst>", content, re.DOTALL)
            if match:
                return match.group(0)
    except Exception:
        pass
    return None


def _restore_workbook_extlst(file_path: Path, extlst: str) -> None:
    """
    Schreibt das gesicherte <extLst>-Element zurück in xl/workbook.xml
    der gespeicherten xlsx-Datei. Verhindert, dass openpyxl die
    AutoSave-Einstellung von Excel deaktiviert.
    """
    tmp_fd, tmp_name = tempfile.mkstemp(suffix=".xlsx", dir=file_path.parent)
    tmp_path = Path(tmp_name)
    try:
        with zipfile.ZipFile(file_path, "r") as zin, \
             zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/workbook.xml":
                    content = data.decode("utf-8")

                    # Namespace-Deklarationen ergänzen, die im extLst referenziert werden
                    def _ensure_namespaces(m: re.Match) -> str:
                        tag = m.group(0)
                        for prefix, uri in _MS_EXT_NAMESPACES:
                            if f"{prefix}:" in extlst and f'xmlns:{prefix}=' not in tag:
                                tag = re.sub(r"(\s*/?>)$", f' xmlns:{prefix}="{uri}"\\1', tag)
                        return tag

                    content = re.sub(r"<workbook\b[^>]*>", _ensure_namespaces, content)

                    # extLst einfügen oder vorhandenes ersetzen
                    if re.search(r"<extLst\b", content):
                        content = re.sub(
                            r"<extLst\b[^>]*>.*?</extLst>",
                            extlst,
                            content,
                            flags=re.DOTALL,
                        )
                    else:
                        content = content.replace("</workbook>", extlst + "</workbook>")

                    data = content.encode("utf-8")
                zout.writestr(item, data)
    except Exception:
        import os
        os.close(tmp_fd)
        tmp_path.unlink(missing_ok=True)
        raise
    else:
        import os
        os.close(tmp_fd)
        shutil.move(str(tmp_path), str(file_path))


def _get_bg_color(cell) -> Optional[str]:
    """Extrahiert die Hintergrundfarbe einer Zelle als Hex-String."""
    fill: PatternFill = cell.fill
    if fill and fill.fgColor:
        color = fill.fgColor
        if color.type == "rgb" and color.rgb != "00000000":
            return color.rgb
    return None


def _get_font_color(cell) -> Optional[str]:
    """Extrahiert die Schriftfarbe einer Zelle."""
    if cell.font and cell.font.color:
        color = cell.font.color
        if color.type == "rgb" and color.rgb != "FF000000":
            return color.rgb
    return None


def _cell_to_model(cell) -> CellInfo:
    """
    Konvertiert eine openpyxl-Zelle in ein CellInfo-Modell.
    MergedCell-Objekte werden als leere Zellen behandelt.
    """
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter

    col_letter = get_column_letter(cell.column)

    if isinstance(cell, MergedCell):
        return CellInfo(
            row=cell.row,
            column=cell.column,
            column_letter=col_letter,
            value=None,
        )

    return CellInfo(
        row=cell.row,
        column=cell.column,
        column_letter=col_letter,
        value=cell.value,
        bg_color=_get_bg_color(cell),
        font_bold=cell.font.bold if cell.font else False,
        font_color=_get_font_color(cell),
    )


# ---------------------------------------------------------------------------
# Öffentliche Klasse
# ---------------------------------------------------------------------------

class ExcelEditor:
    """
    Haupt-Editor-Klasse (openpyxl-Backend).

    Verwendet openpyxl zum Lesen und Schreiben – kein laufendes Excel nötig.
    Formatierung (Farben, Schrift, Rahmen, Formeln) wird vollständig erhalten.

    Verwendung:
        with ExcelEditor(config) as editor:
            editor.move_row_after("1010", "1026")
            editor.save()
    """

    def __init__(self, config: ExcelReadConfig) -> None:
        self.config = config
        # AutoSave-Erweiterung vor dem Laden sichern, da openpyxl sie beim
        # Speichern nicht erhält (x15:workbookPr autoSave="1").
        self._workbook_extlst: Optional[str] = _read_workbook_extlst(config.file_path)
        # data_only=False -> Formeln werden als Formelstring erhalten
        self._workbook = openpyxl.load_workbook(
            config.file_path, data_only=False
        )
        self._worksheet = self._get_worksheet()

    # ------------------------------------------------------------------
    # Sheet-Zugriff
    # ------------------------------------------------------------------

    def _get_worksheet(self):
        """Gibt das konfigurierte Worksheet zurück."""
        if self.config.sheet_name:
            if self.config.sheet_name not in self._workbook.sheetnames:
                available = self._workbook.sheetnames
                raise ValueError(
                    f"Sheet '{self.config.sheet_name}' nicht gefunden. "
                    f"Verfügbar: {available}"
                )
            return self._workbook[self.config.sheet_name]
        return self._workbook.active

    def get_sheet_names(self) -> List[str]:
        """Gibt alle Sheet-Namen zurück."""
        return self._workbook.sheetnames

    def get_sheet_info(self) -> SheetInfo:
        """Gibt Metadaten des aktuellen Sheets zurück."""
        ws = self._worksheet
        header_row = self.config.header_row
        headers: Dict[int, str] = {}
        for cell in ws[header_row]:
            headers[cell.column] = (
                str(cell.value) if cell.value is not None
                else f"Spalte_{cell.column}"
            )
        return SheetInfo(
            name=ws.title,
            max_row=ws.max_row,
            max_column=ws.max_column,
            headers=headers,
            header_row=header_row,
        )

    # ------------------------------------------------------------------
    # Daten lesen
    # ------------------------------------------------------------------

    def get_rows(
        self,
        min_row: Optional[int] = None,
        max_row: Optional[int] = None,
        skip_empty: bool = True,
    ) -> List[RowData]:
        """
        Liest alle Zeilen nach der Header-Zeile aus.

        Args:
            min_row: Erste Zeile (Standard: header_row + 1)
            max_row: Letzte Zeile (Standard: bis Ende)
            skip_empty: Leere Zeilen überspringen
        """
        if min_row is None:
            min_row = self.config.header_row + 1
        ws = self._worksheet
        rows: List[RowData] = []

        for row in ws.iter_rows(min_row=min_row, max_row=max_row):
            cells = [_cell_to_model(cell) for cell in row]
            if skip_empty and all(c.value is None for c in cells):
                continue
            rows.append(RowData(row_index=row[0].row, cells=cells))

        return rows

    def get_row(self, row_index: int) -> Optional[RowData]:
        """Gibt eine einzelne Zeile anhand des Index zurück."""
        ws = self._worksheet
        row = ws[row_index]
        cells = [_cell_to_model(cell) for cell in row]
        if all(c.value is None for c in cells):
            return None
        return RowData(row_index=row_index, cells=cells)

    # ------------------------------------------------------------------
    # Daten schreiben
    # ------------------------------------------------------------------

    def edit_cell(self, row: int, column: int, new_value: Any) -> None:
        """
        Bearbeitet eine einzelne Zelle. Formatierung bleibt erhalten.

        Args:
            row: Zeilenindex (1-basiert)
            column: Spaltenindex (1-basiert)
            new_value: Neuer Wert der Zelle
        """
        ws = self._worksheet
        cell = ws.cell(row=row, column=column)
        old_font = copy.copy(cell.font)
        old_fill = copy.copy(cell.fill)
        old_border = copy.copy(cell.border)
        old_alignment = copy.copy(cell.alignment)
        old_number_format = cell.number_format

        cell.value = new_value

        cell.font = old_font
        cell.fill = old_fill
        cell.border = old_border
        cell.alignment = old_alignment
        cell.number_format = old_number_format

    def edit_row(self, row_index: int, updates: Dict[int, Any]) -> None:
        """
        Bearbeitet mehrere Zellen einer Zeile auf einmal.

        Args:
            row_index: Zeilenindex (1-basiert)
            updates: Dict mit {Spaltenindex: neuer_Wert}
        """
        for column_index, new_value in updates.items():
            self.edit_cell(row=row_index, column=column_index, new_value=new_value)

    # ------------------------------------------------------------------
    # Zeile verschieben
    # ------------------------------------------------------------------

    def _find_no_column(self) -> int:
        """Findet den Spaltenindex der 'No'-Spalte."""
        info = self.get_sheet_info()
        for idx, name in info.headers.items():
            if name.strip().lower() == "no":
                return idx
        raise ValueError(
            "Spalte 'No' nicht im Sheet gefunden. "
            f"Verfügbare Spalten: {list(info.headers.values())}"
        )

    def _find_row_by_no(self, no_value: Any) -> int:
        """
        Sucht den Zeilenindex der Zeile mit dem angegebenen 'No'-Wert.
        Wirft ValueError wenn nicht gefunden.
        """
        from openpyxl.cell.cell import MergedCell

        no_col = self._find_no_column()
        ws = self._worksheet
        min_row = self.config.header_row + 1

        for row in ws.iter_rows(min_row=min_row):
            cell = row[no_col - 1]
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None and str(cell.value) == str(no_value):
                return row[0].row

        raise ValueError(f"Zeile mit No='{no_value}' nicht gefunden.")

    def _copy_row_data(self, row_index: int) -> List[Dict]:
        """Kopiert alle Zell-Daten (Wert + Format) einer Zeile."""
        from openpyxl.cell.cell import MergedCell

        ws = self._worksheet
        copied: List[Dict] = []

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_index, column=col_idx)
            if isinstance(cell, MergedCell):
                copied.append({"col": col_idx, "merged": True})
                continue
            copied.append({
                "col": col_idx,
                "merged": False,
                "value": cell.value,
                "font": copy.copy(cell.font),
                "fill": copy.copy(cell.fill),
                "border": copy.copy(cell.border),
                "alignment": copy.copy(cell.alignment),
                "number_format": cell.number_format,
            })

        return copied

    def _paste_row_data(self, row_index: int, copied: List[Dict]) -> None:
        """Schreibt zuvor kopierte Zell-Daten in eine Zeile."""
        ws = self._worksheet
        for cell_data in copied:
            if cell_data.get("merged"):
                continue
            cell = ws.cell(row=row_index, column=cell_data["col"])
            cell.value = cell_data["value"]
            cell.font = cell_data["font"]
            cell.fill = cell_data["fill"]
            cell.border = cell_data["border"]
            cell.alignment = cell_data["alignment"]
            cell.number_format = cell_data["number_format"]

    def _find_insert_position_for_no(self, target_no: Any) -> int:
        """
        Gibt den Zeilenindex der letzten Zeile zurück, deren No-Wert < target_no ist.
        Wird verwendet, wenn after_no im Sheet nicht existiert.
        Gibt den Index der letzten Datenzeile zurück, wenn alle No-Werte >= target_no.
        """
        from openpyxl.cell.cell import MergedCell

        try:
            target_int = int(target_no)
        except (ValueError, TypeError):
            raise ValueError(f"No='{target_no}' ist kein ganzzahliger Wert.")

        no_col = self._find_no_column()
        ws = self._worksheet
        min_row = self.config.header_row + 1

        last_smaller_idx: Optional[int] = None
        last_data_idx: int = min_row

        for row in ws.iter_rows(min_row=min_row):
            cell = row[no_col - 1]
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            last_data_idx = row[0].row
            try:
                no_int = int(cell.value)
            except (ValueError, TypeError):
                continue
            if no_int < target_int:
                last_smaller_idx = row[0].row

        return last_smaller_idx if last_smaller_idx is not None else last_data_idx

    def move_row_after(self, source_no: Any, after_no: Any) -> int:
        """
        Verschiebt die Zeile mit No=source_no direkt NACH die Zeile mit No=after_no.

        Existiert after_no im Sheet:
            new_no = int((after_no + next_below_no) / 2)
        Existiert after_no NICHT:
            Zeile wird an der passenden Stelle eingefügt und erhält new_no = after_no.

        Bricht mit ValueError ab wenn:
          - source_no nicht existiert
          - Kein ganzzahliger Mittelwert möglich (nur wenn after_no existiert)
          - Das berechnete new_no bereits als No existiert
          - after_no und source_no identisch sind

        Returns:
            new_no: Der tatsächlich vergebene No-Wert
        """
        from openpyxl.cell.cell import MergedCell

        no_col = self._find_no_column()
        ws = self._worksheet

        source_idx = self._find_row_by_no(source_no)

        # after_no muss nicht existieren – dann direkt als neues No verwenden
        try:
            after_idx = self._find_row_by_no(after_no)
            after_no_exists = True
        except ValueError:
            after_no_exists = False
            after_idx = self._find_insert_position_for_no(after_no)

        # "Identisch"-Fehler nur wenn after_no tatsächlich existiert
        if after_no_exists and source_idx == after_idx:
            raise ValueError("Quell- und Zielzeile sind identisch.")

        try:
            after_no_int = int(after_no)
        except (ValueError, TypeError):
            raise ValueError(f"No='{after_no}' ist kein ganzzahliger Wert.")

        if after_no_exists:
            # Nächste Zeile unter after_idx finden (source überspringen)
            below_no_int: Optional[int] = None
            for row in ws.iter_rows(min_row=after_idx + 1):
                if row[0].row == source_idx:
                    continue
                cell = row[no_col - 1]
                if isinstance(cell, MergedCell) or cell.value is None:
                    continue
                try:
                    below_no_int = int(cell.value)
                    break
                except (ValueError, TypeError):
                    continue

            # Neues No berechnen
            if below_no_int is None:
                new_no = after_no_int + 10
            else:
                raw_mid = (after_no_int + below_no_int) / 2
                new_no = int(raw_mid)
                if new_no <= after_no_int:
                    raise ValueError(
                        f"Kein ganzzahliger Mittelwert möglich zwischen "
                        f"No={after_no_int} und No={below_no_int} "
                        f"(Mittelwert wäre {raw_mid}). "
                        f"Bitte Nummernlücke vergrößern."
                    )

            # Prüfen ob new_no bereits existiert
            try:
                self._find_row_by_no(new_no)
                raise ValueError(
                    f"Berechnetes No={new_no} ist bereits vergeben. "
                    f"Bitte Nummernlücke zwischen {after_no_int} und "
                    f"{below_no_int} vergrößern."
                )
            except ValueError as e:
                if "bereits vergeben" in str(e):
                    raise
        else:
            # after_no existiert nicht → direkt als neues No vergeben
            new_no = after_no_int

        # Source-Zeile sichern und No-Wert überschreiben
        copied_cells = self._copy_row_data(source_idx)
        source_height = ws.row_dimensions[source_idx].height

        for cell_data in copied_cells:
            if not cell_data.get("merged") and cell_data["col"] == no_col:
                cell_data["value"] = new_no
                break

        # Verschieben
        if source_idx == after_idx:
            # after_no nicht vorhanden und source steht bereits an der richtigen Position
            # → nur No-Wert umbenennen, keine Zeile verschieben
            self._paste_row_data(source_idx, copied_cells)
        elif source_idx < after_idx:
            ws.delete_rows(source_idx)
            adjusted_after = after_idx - 1
            ws.insert_rows(adjusted_after + 1)
            self._paste_row_data(adjusted_after + 1, copied_cells)
            if source_height:
                ws.row_dimensions[adjusted_after + 1].height = source_height
        else:
            ws.insert_rows(after_idx + 1)
            self._paste_row_data(after_idx + 1, copied_cells)
            if source_height:
                ws.row_dimensions[after_idx + 1].height = source_height
            ws.delete_rows(source_idx + 1)

        return new_no

    # Rückwärtskompatibilität
    def renumber_and_move_row(self, source_no: Any, after_no: Any) -> tuple:
        new_no = self.move_row_after(source_no, after_no)
        return new_no, False

    def move_row_by_no(self, source_no: Any, after_no: Any) -> None:
        self.move_row_after(source_no, after_no)

    # ------------------------------------------------------------------
    # Speichern
    # ------------------------------------------------------------------

    def save(self, output_path: Optional[Path] = None) -> Path:
        """
        Speichert die Datei.

        Args:
            output_path: Zielpfad (Standard: überschreibt Originaldatei)

        Returns:
            Pfad zur gespeicherten Datei
        """
        target = output_path or self.config.file_path
        self._workbook.save(target)
        # AutoSave-Erweiterung wiederherstellen, die openpyxl entfernt hat
        if self._workbook_extlst:
            _restore_workbook_extlst(target, self._workbook_extlst)
        return target

    def close(self) -> None:
        """Schließt die Arbeitsmappe."""
        self._workbook.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
