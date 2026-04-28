"""
Tests für ExcelEditor: edit_cell, move_row_after, AutoSave-Erhalt.
"""
import shutil
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from excel_editor.editor import ExcelEditor, _read_workbook_extlst, _restore_workbook_extlst
from excel_editor.models import ExcelReadConfig


def _make_workbook(path: Path) -> None:
    """Erstellt eine minimale Test-Excel-Datei mit No-Spalte."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "No"
    ws["B1"] = "Beschreibung"
    ws["A2"] = 1000
    ws["B2"] = "Erste Zeile"
    ws["A3"] = 1010
    ws["B3"] = "Zweite Zeile"
    ws["A4"] = 1020
    ws["B4"] = "Dritte Zeile"
    wb.save(path)


def _add_extlst(path: Path) -> None:
    """Fügt ein <extLst> mit autoSave='1' in workbook.xml ein."""
    EXTLST = (
        '<extLst>'
        '<ext uri="{140A7094-0E35-4892-8432-C4D2E57EDEB5}"'
        ' xmlns:x15="http://schemas.microsoft.com/office/spreadsheetml/2010/11/main">'
        '<x15:workbookPr autoSave="1"/>'
        '</ext>'
        '</extLst>'
    )
    tmp = path.parent / (path.name + ".tmp")
    with zipfile.ZipFile(path, "r") as zin, \
         zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/workbook.xml":
                content = data.decode("utf-8")
                content = content.replace("</workbook>", EXTLST + "</workbook>")
                data = content.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(str(tmp), str(path))


class TestEditCell(unittest.TestCase):

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.xlsx = self.tmp / "test.xlsx"
        _make_workbook(self.xlsx)

    def tearDown(self):
        shutil.rmtree(self.tmp)

    def test_edit_cell_value(self):
        config = ExcelReadConfig(file_path=self.xlsx)
        with ExcelEditor(config) as ed:
            ed.edit_cell(row=2, column=2, new_value="Geändert")
            ed.save()

        config2 = ExcelReadConfig(file_path=self.xlsx)
        with ExcelEditor(config2) as ed2:
            row = ed2.get_row(2)
        self.assertEqual(row.get_value(2), "Geändert")

    def test_edit_cell_preserves_formatting(self):
        """Formatierung (Fett, Farbe) bleibt nach edit_cell erhalten."""
        from openpyxl import load_workbook as lw
        wb = lw(self.xlsx)
        ws = wb.active
        ws["B2"].font = Font(bold=True)
        ws["B2"].fill = PatternFill("solid", fgColor="FFFF00")
        wb.save(self.xlsx)

        config = ExcelReadConfig(file_path=self.xlsx)
        with ExcelEditor(config) as ed:
            ed.edit_cell(row=2, column=2, new_value="Neu")
            ed.save()

        wb2 = lw(self.xlsx)
        cell = wb2.active["B2"]
        self.assertTrue(cell.font.bold)
        self.assertIn("FFFF00", cell.fill.fgColor.rgb)


class TestMoveRow(unittest.TestCase):

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.xlsx = self.tmp / "test.xlsx"
        _make_workbook(self.xlsx)

    def tearDown(self):
        shutil.rmtree(self.tmp)

    def _load(self):
        config = ExcelReadConfig(file_path=self.xlsx)
        return ExcelEditor(config)

    def test_move_row_after_calculates_new_no(self):
        """No zwischen 1000 und 1010 → erwartet 1005."""
        with self._load() as ed:
            new_no = ed.move_row_after("1020", "1000")
            ed.save()
        self.assertEqual(new_no, 1005)

    def test_move_row_after_appends_if_last(self):
        """Verschoben ans Ende → No = after_no + 10."""
        with self._load() as ed:
            new_no = ed.move_row_after("1000", "1020")
            ed.save()
        self.assertEqual(new_no, 1030)

    def test_move_row_after_identical_raises(self):
        with self._load() as ed:
            with self.assertRaises(ValueError):
                ed.move_row_after("1000", "1000")

    def test_move_row_after_unknown_no_raises(self):
        with self._load() as ed:
            with self.assertRaises(ValueError):
                ed.move_row_after("9999", "1000")


class TestAutoSave(unittest.TestCase):

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.xlsx = self.tmp / "test.xlsx"
        _make_workbook(self.xlsx)
        _add_extlst(self.xlsx)

    def tearDown(self):
        shutil.rmtree(self.tmp)

    def test_autosave_preserved_after_save(self):
        """autoSave='1' bleibt nach ExcelEditor.save() erhalten."""
        extlst_before = _read_workbook_extlst(self.xlsx)
        self.assertIsNotNone(extlst_before)
        self.assertIn('autoSave="1"', extlst_before)

        config = ExcelReadConfig(file_path=self.xlsx)
        with ExcelEditor(config) as ed:
            ed.edit_cell(row=2, column=2, new_value="test")
            ed.save()

        extlst_after = _read_workbook_extlst(self.xlsx)
        self.assertIsNotNone(extlst_after)
        self.assertIn('autoSave="1"', extlst_after)

    def test_autosave_preserved_after_move_row(self):
        """autoSave='1' bleibt nach move_row_after + save() erhalten."""
        config = ExcelReadConfig(file_path=self.xlsx)
        with ExcelEditor(config) as ed:
            ed.move_row_after("1020", "1000")
            ed.save()

        extlst_after = _read_workbook_extlst(self.xlsx)
        self.assertIsNotNone(extlst_after)
        self.assertIn('autoSave="1"', extlst_after)


class TestRollback(unittest.TestCase):

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.xlsx = self.tmp / "test.xlsx"
        _make_workbook(self.xlsx)

    def tearDown(self):
        shutil.rmtree(self.tmp)

    def test_rollback_does_not_save(self):
        """--rollback: move_row wird berechnet, Datei bleibt unveraendert."""
        import zipfile as zf

        # Originalen Inhalt von workbook.xml sichern
        with zf.ZipFile(self.xlsx, "r") as z:
            original_wb_xml = z.read("xl/workbook.xml")

        config = ExcelReadConfig(file_path=self.xlsx)
        with ExcelEditor(config) as ed:
            ed.move_row_after("1020", "1000")
            # rollback = NICHT speichern

        # Datei darf sich nicht veraendert haben
        with zf.ZipFile(self.xlsx, "r") as z:
            after_wb_xml = z.read("xl/workbook.xml")

        self.assertEqual(original_wb_xml, after_wb_xml)

    def test_save_does_change_file(self):
        """Kontrolltest: save() veraendert die Datei tatsaechlich."""
        import zipfile as zf

        with zf.ZipFile(self.xlsx, "r") as z:
            original = z.read("xl/worksheets/sheet1.xml")

        config = ExcelReadConfig(file_path=self.xlsx)
        with ExcelEditor(config) as ed:
            ed.edit_cell(row=2, column=2, new_value="GEAENDERT")
            ed.save()

        with zf.ZipFile(self.xlsx, "r") as z:
            after = z.read("xl/worksheets/sheet1.xml")

        self.assertNotEqual(original, after)


class TestSheetRouting(unittest.TestCase):
    """Prueft dass master_sheet / sid_sheet korrekt an _process_single_file
    weitergegeben werden."""

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.xlsx = self.tmp / "test.xlsx"
        # Zwei Sheets erstellen
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "MasterSheet"
        ws1["A1"] = "No"
        ws1["A2"] = 1000
        ws2 = wb.create_sheet("SIDSheet")
        ws2["A1"] = "No"
        ws2["A2"] = 2000
        wb.save(self.xlsx)

    def tearDown(self):
        shutil.rmtree(self.tmp)

    def test_correct_sheet_is_loaded(self):
        """Sheet-Name wird korrekt gesetzt."""
        config = ExcelReadConfig(file_path=self.xlsx, sheet_name="SIDSheet")
        with ExcelEditor(config) as ed:
            row = ed.get_row(2)
        self.assertEqual(row.get_value(1), 2000)

    def test_wrong_sheet_raises(self):
        """Unbekannter Sheet-Name wirft ValueError."""
        from pydantic import ValidationError as PydanticValidationError
        config = ExcelReadConfig(file_path=self.xlsx)
        with ExcelEditor(config) as ed:
            with self.assertRaises(ValueError):
                config.sheet_name = "GibtEsNicht"
                ed._worksheet = ed._get_worksheet()


if __name__ == "__main__":
    unittest.main()